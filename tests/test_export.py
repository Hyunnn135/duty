"""엑셀(.xlsx) 내보내기 엔드포인트·빌더 테스트."""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient

from app.export_xlsx import ExportRequest, build_xlsx


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("DUTY_DB", str(tmp_path / "test.db"))
    monkeypatch.setenv("DUTY_SECRET", "test-secret")
    from app.main import app

    return TestClient(app)


def _sample() -> ExportRequest:
    return ExportRequest(
        year=2026, month=8, num_days=31, holidays=[15, 17],
        teams={"김간호": 1, "이간호": 2},
        schedules=[
            {"name": "김간호", "labels": ["D"] * 31, "counts": {"D": 31}},
            {"name": "이간호", "labels": ["N"] * 15 + ["O"] * 16, "counts": {"N": 15, "O": 16}},
        ],
    )


def test_build_xlsx_returns_valid_workbook():
    data = build_xlsx(_sample())
    assert data[:2] == b"PK"  # xlsx = zip
    # openpyxl로 되읽어 값 검증
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    # 헤더행(4) 아래에 간호사 이름이 들어있는지
    names = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert "김간호" in names and "이간호" in names
    assert any(str(v).startswith("Σ") for v in names)  # 합계행 존재


def test_export_endpoint_requires_auth_and_returns_xlsx(client):
    payload = _sample().model_dump()
    # 미인증 → 401
    assert client.post("/api/schedule/export.xlsx", json=payload).status_code == 401
    # 인증 → 200 + xlsx
    tok = client.post("/api/auth/register", json={
        "email": "boss@d.kr", "name": "파트장", "password": "password123", "ward": "61",
    }).json()["token"]
    r = client.post("/api/schedule/export.xlsx", json=payload,
                    headers={"Authorization": f"Bearer {tok}"})
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"


def test_export_rejects_out_of_range_year_month(client):
    tok = client.post("/api/auth/register", json={
        "email": "b2@d.kr", "name": "파트장", "password": "password123", "ward": "62",
    }).json()["token"]
    payload = _sample().model_dump()
    payload["year"], payload["month"] = 5, 13
    r = client.post("/api/schedule/export.xlsx", json=payload,
                    headers={"Authorization": f"Bearer {tok}"})
    assert r.status_code == 422


def test_formula_injection_neutralized():
    """'='로 시작하는 사용자 문자열이 수식이 아닌 텍스트로 저장된다(주입 차단)."""
    req = ExportRequest(
        year=2026, month=8, num_days=3, teams={},
        title='=HYPERLINK("http://evil")', subtitle="=1+1",
        schedules=[{"name": "=CMD()", "labels": ["=2+3", "D", "O"], "counts": {}}],
    )
    data = build_xlsx(req)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    for row in ws.iter_rows():
        for c in row:
            assert c.data_type != "f", f"수식으로 저장됨: {c.coordinate}={c.value!r}"
    assert ws["A1"].value == '=HYPERLINK("http://evil")'
