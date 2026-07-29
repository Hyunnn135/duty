"""근무표 → 엑셀(.xlsx) 변환.

프론트엔드가 보낸 근무표(이름·라벨·카운트·팀·연월·공휴일)를 받아 서식 있는 워크북
바이트를 만든다. 스타일은 배포 전 검증한 스냅샷 서식을 따른다(주말/공휴일 음영, 팀 구분선,
교대별 일일 합계 Σ). 카운트는 스냅샷이므로 정적 값으로 기록한다(수식 없음).
"""
from __future__ import annotations

import io
from datetime import date

from pydantic import BaseModel, Field

_COLOR = {"D": "2563EB", "E": "D97706", "N": "7C3AED", "M": "0E8FA0", "O": "94A3B8"}
_WD = ["월", "화", "수", "목", "금", "토", "일"]


class ExportSchedule(BaseModel):
    name: str
    labels: list[str]
    counts: dict[str, int] = Field(default_factory=dict)


class ExportRequest(BaseModel):
    year: int | None = None
    month: int | None = None
    num_days: int = Field(31, ge=1, le=62)
    holidays: list[int] = Field(default_factory=list)
    teams: dict[str, int] = Field(default_factory=dict)
    schedules: list[ExportSchedule]
    title: str | None = None
    subtitle: str | None = None


def _norm(lab: str) -> str:
    return lab if lab in ("D", "E", "N", "M", "O") else ("M" if lab in ("M/D", "M/E") else "O")


def build_xlsx(req: ExportRequest) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter as col

    nd = req.num_days
    hol = set(req.holidays)
    # 1일의 요일(월=0). year/month 없으면 토요일(5) 가정(표시는 근사).
    fw = date(req.year, req.month, 1).weekday() if (req.year and req.month) else 5

    def wknd(day0: int) -> bool:
        return (fw + day0) % 7 >= 5 or (day0 + 1) in hol

    teams = req.teams
    rows = sorted(req.schedules, key=lambda s: (teams.get(s.name, 9), s.name))

    wb = Workbook()
    ws = wb.active
    ws.title = "근무표"
    AR = "Arial"
    thin = Side(style="thin", color="D9E1EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wkfill = PatternFill("solid", fgColor="FFF7ED")
    hdrfill = PatternFill("solid", fgColor="F1F5F9")
    sepfill = PatternFill("solid", fgColor="EEF2F7")
    center = Alignment(horizontal="center", vertical="center")

    period = f"{req.year}년 {req.month}월 " if (req.year and req.month) else ""
    ws["A1"] = req.title or f"{period}근무표"
    ws["A1"].font = Font(name=AR, size=13, bold=True)
    if req.subtitle:
        ws["A2"] = req.subtitle
        ws["A2"].font = Font(name=AR, size=9, color="475569")

    HR = 4
    ws.cell(HR, 1, "간호사").font = Font(name=AR, bold=True)
    ws.cell(HR, 2, "팀").font = Font(name=AR, bold=True)
    for day in range(1, nd + 1):
        c = ws.cell(HR, 2 + day, day)
        c.font = Font(name=AR, bold=True, size=9); c.alignment = center; c.fill = hdrfill; c.border = border
        w = ws.cell(HR + 1, 2 + day, _WD[(fw + day - 1) % 7])
        w.font = Font(name=AR, size=8, color="94A3B8"); w.alignment = center
        if wknd(day - 1):
            c.fill = wkfill; w.fill = wkfill
        c.border = border; w.border = border
    ws.cell(HR + 1, 1, "").border = border
    cnt_cols = {"D": 3 + nd, "E": 4 + nd, "N": 5 + nd, "M": 6 + nd, "O": 7 + nd}
    for lab, cc in cnt_cols.items():
        h = ws.cell(HR, cc, lab)
        h.font = Font(name=AR, bold=True); h.alignment = center; h.fill = hdrfill; h.border = border
        ws.cell(HR + 1, cc, "").border = border

    r = HR + 2
    prev = None
    for s in rows:
        t = teams.get(s.name, 9)
        nm = ws.cell(r, 1, s.name); nm.font = Font(name=AR, bold=True, size=10); nm.border = border
        tc = ws.cell(r, 2, f"{t}팀"); tc.font = Font(name=AR, size=9, color="3730A3"); tc.alignment = center; tc.border = border
        if prev is not None and t != prev:
            for cc in range(1, 8 + nd):
                cur = ws.cell(r, cc)
                cur.border = Border(left=thin, right=thin, bottom=thin, top=Side(style="medium", color="CBD5E1"))
        prev = t
        for day in range(nd):
            lab = s.labels[day] if day < len(s.labels) else "O"
            base = _norm(lab)
            cell = ws.cell(r, 3 + day, lab); cell.alignment = center; cell.border = border
            cell.font = Font(name=AR, bold=True, size=10, color=_COLOR.get(base, "BE185D"))
            if wknd(day):
                cell.fill = wkfill
        for lab, cc in cnt_cols.items():
            f = ws.cell(r, cc, s.counts.get(lab, 0))
            f.alignment = center; f.font = Font(name=AR, bold=True); f.fill = hdrfill; f.border = border
        r += 1

    # Σ 일별 합계(D/E/N/M)
    for lab in ("D", "E", "N", "M"):
        lc = ws.cell(r, 1, f"Σ {lab}"); lc.font = Font(name=AR, bold=True); lc.fill = sepfill; lc.border = border
        ws.cell(r, 2, "").fill = sepfill; ws.cell(r, 2).border = border
        for day in range(nd):
            cnt = sum(1 for s in rows if day < len(s.labels) and _norm(s.labels[day]) == lab)
            f = ws.cell(r, 3 + day, cnt)
            f.alignment = center; f.font = Font(name=AR, bold=True, size=9); f.fill = sepfill; f.border = border
        for cc in cnt_cols.values():
            ws.cell(r, cc, "").fill = sepfill; ws.cell(r, cc).border = border
        r += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 5
    for day in range(1, nd + 1):
        ws.column_dimensions[col(2 + day)].width = 4.2
    for cc in cnt_cols.values():
        ws.column_dimensions[col(cc)].width = 4.5
    ws.freeze_panes = ws.cell(HR + 2, 3)
    ws.row_dimensions[HR].height = 15
    ws.row_dimensions[HR + 1].height = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
